import os
import pandas as pd
import matplotlib
matplotlib.use("Agg")  # Headless mode
import matplotlib.pyplot as plt
from evo.tools import file_interface
from evo.core import sync, metrics
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# === Paths ===
gt_dir = "/mnt/d/Datasets/SLAM_Datasets/vKITTI/vkitti_1.3.1_extrinsicsgt"
est_dir = "droid_vkitti_experiments"
output_dir = "droid_vkitti_experiments/eval_results"
os.makedirs(output_dir, exist_ok=True)

sequences = ["0001", "0002", "0006", "0018", "0020"]
weathers = ["morning", "sunset", "clone", "fog", "overcast", "rain"]

results = []

print("Starting evaluations...\n")

for seq in sequences:
    for weather in weathers:
        tag = f"{seq}-{weather}"
        print(f"▶ Evaluating {tag}...")

        gt_path = os.path.join(gt_dir, f"{seq}_{weather}.tum")
        est_path = os.path.join(est_dir, f"droid_vKITTI_{seq}_{weather}_traj.txt")

        if not (os.path.exists(gt_path) and os.path.exists(est_path)):
            print(f"⚠ Missing files for {tag}")
            continue

        try:
            traj_ref = file_interface.read_tum_trajectory_file(gt_path)
            traj_est = file_interface.read_tum_trajectory_file(est_path)

            traj_ref, traj_est = sync.associate_trajectories(traj_ref, traj_est)
            traj_est.align(traj_ref, correct_scale=True)

            # Compute APE
            ape_metric = metrics.APE(metrics.PoseRelation.translation_part)
            ape_metric.process_data((traj_ref, traj_est))
            ape_stats = ape_metric.get_all_statistics()
            ape_array = np.array(ape_metric.error)

            # === Custom APE colormap plot (default evo_ape style) ===
            fig, ax = plt.subplots(figsize=(6, 5))
            sc = ax.scatter(
                traj_est.positions_xyz[:, 0],
                traj_est.positions_xyz[:, 2],
                c=ape_array,
                cmap="plasma",
                s=10,
                label="Estimated (colored by APE)"
            )
            ax.plot(
                traj_ref.positions_xyz[:, 0],
                traj_ref.positions_xyz[:, 2],
                "--",
                color="black",
                linewidth=1,
                alpha=0.6,
                label="Ground Truth"
            )

            ax.set_title(f"{seq}-{weather} APE (color by error)")
            ax.set_xlabel("x [m]")
            ax.set_ylabel("z [m]")
            ax.legend(loc="upper right")
            cbar = plt.colorbar(sc, ax=ax)
            cbar.set_label("APE [m]")
            plt.tight_layout()

            plot_path = os.path.join(output_dir, f"{tag}_ape_plot.png")
            plt.savefig(plot_path, dpi=200, bbox_inches="tight")
            plt.close(fig)

            results.append({
                "Sequence": seq,
                "Weather": weather,
                "Mean APE (m)": ape_stats["mean"],
                "Median APE (m)": ape_stats["median"],
                "Std APE (m)": ape_stats["std"],
                "RMSE APE (m)": ape_stats["rmse"],
                "Min APE (m)": ape_stats["min"],
                "Max APE (m)": ape_stats["max"],
                "Plot Path": plot_path
            })

        except Exception as e:
            print(f"⚠ Error during evaluation of {tag}: {e}")
            continue

# === Export to Excel with thumbnails ===
if results:
    print("\n📊 Saving to Excel with thumbnails...")

    excel_path = os.path.join(output_dir, "vkitti_ape_results.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "APE Results"

    headers = ["Sequence", "Weather", "Mean APE (m)", "Median APE (m)",
               "Std APE (m)", "RMSE APE (m)", "Min APE (m)", "Max APE (m)", "Trajectory Thumbnail"]
    ws.append(headers)

    # Adjust column widths for all columns
    col_widths = [12, 12, 14, 14, 12, 14, 12, 12, 25]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Start filling data
    for i, r in enumerate(results, start=2):
        ws.append([r["Sequence"], r["Weather"], r["Mean APE (m)"], r["Median APE (m)"],
                   r["Std APE (m)"], r["RMSE APE (m)"], r["Min APE (m)"], r["Max APE (m)"], ""])
        # Add thumbnail
        img = XLImage(r["Plot Path"])
        img.width = 150  # Thumbnail width in pixels
        img.height = 150  # Thumbnail height in pixels
        ws.add_image(img, f"I{i}")
        # Set row height to fit thumbnail
        ws.row_dimensions[i].height = 110  # approximately fits 150px image

    wb.save(excel_path)
    print(f"✅ Evaluation complete. Results with thumbnails saved to:\n{excel_path}")

else:
    print("\n❌ No valid evaluations — check timestamps or trajectory alignment.")